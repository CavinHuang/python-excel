from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle
from copy import copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

alignment = Alignment(wrapText=True, shrinkToFit=False, wrap_text=True, vertical='center', horizontal='center')

# def copy_style(source_style, target_cell):
#     """
#     复制源样式到目标单元格
#     """
#     new_style = NamedStyle(name=source_style.name)
#     for key, value in source_style.__dict__.items():
#         setattr(new_style, key, value)
#     target_cell.style = new_style

def copy_style(source_style, target_cell):
    target_cell.font = copy(source_style.font)
    target_cell.border = copy(source_style.border)
    target_cell.fill = copy(source_style.fill)
    target_cell.number_format = copy(source_style.number_format)
    target_cell.protection = copy(source_style.protection)
    target_cell.alignment = copy(source_style.alignment)

def xlsx_sheet_copy(src_file_sheet, tag_file_sheet):  # 跨xlsx复制sheet
    # 跨xlsx文件复制源文件xlsx中指定的sheet
    # 保留所有格式，以及行高列宽，视觉效果几乎一致
    for row in src_file_sheet:
        tag_row = []  # 用于存储新行的数据
        for cell in row:  # 复制数据
            tag_cell = tag_file_sheet[cell.coordinate]
            tag_cell.value = cell.value

            if cell.has_style:  # 复制样式
                tag_cell.font = copy(cell.font)
                tag_cell.border = copy(cell.border)
                tag_cell.fill = copy(cell.fill)
                tag_cell.number_format = copy(cell.number_format)
                tag_cell.protection = copy(cell.protection)
                tag_cell.alignment = alignment
            
            tag_row.append(tag_cell)

        # 设置新行高为默认值（可根据需求调整）
        tag_row_dimension = tag_file_sheet.row_dimensions[row[0].row]
        tag_row_dimension.height = None

        tag_file_sheet.append(tag_row)  # 添加新行

    wm = list(zip(src_file_sheet.merged_cells))  # 开始处理合并单元格
    if len(wm) > 0:  # 检测源xlsx中合并的单元格
        for i in range(0, len(wm)):
            cell2 = (str(wm[i]).replace("(<MergedCellRange ", "").replace(">,)", ""))  # 获取合并单元格的范围
            tag_file_sheet.merge_cells(cell2)  # 合并单元格

    # 开始处理行高
    for i, row_dimension in src_file_sheet.row_dimensions.items():
        tag_row_dimension = tag_file_sheet.row_dimensions[i]
        tag_row_dimension.height = row_dimension.height

    # 开始处理列宽
    for i in range(1, src_file_sheet.max_column + 1):
        col_letter = get_column_letter(i)
        src_col_dimension = src_file_sheet.column_dimensions[col_letter]
        tag_col_dimension = tag_file_sheet.column_dimensions[col_letter]
        tag_col_dimension.width = src_col_dimension.width