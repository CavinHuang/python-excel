from openpyxl import load_workbook, Workbook
import os
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from copy_worksheet import xlsx_sheet_copy
from openpyxl.styles import Alignment, Font
from copy import copy


# 设置单元格的对齐方式，使内容溢出时隐藏
alignment = Alignment(wrapText=True, shrinkToFit=False, wrap_text=True, vertical='center', horizontal='center')

def copy_images(src_sheet, dest_sheet):
    print(src_sheet)
    for img in src_sheet._images:
        image = Image(img.image)
        print('tupina', img.anchor)
        dest_sheet.add_image(image, img.anchor)

def process_excel(file_path):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        new_workbook = Workbook()
        copy_sheet = new_workbook.active
        # new_workbook.active = sheet.copy()
        copy_sheet.title = sheet.title

        # old_sheet = new_workbook.create_sheet(title=sheet.title)
        new_sheet = new_workbook.create_sheet(title="sheet 1")

        # 复制旧表格内容到第一个sheet
        # for row in sheet.iter_rows(values_only=True):
        #     old_sheet.append(row)
        # xlsx_sheet_copy(sheet, copy_sheet)

        # 获取合并单元格的范围
        merged_ranges = sheet.merged_cells.ranges

        # 遍历每一行数据
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            # print('🚀 ~ file: data_processor.py:22 ~ row_idx:', row_idx, row)
            new_row = []
            in_merged_range = False  # 标记当前单元格是否在合并单元格范围内
            merged_range = None  # 初始化合并单元格范围

            # 设置行高为20
            new_sheet.row_dimensions[row_idx].height = 20
            for col_idx, cell in enumerate(row, start=1):
                # current_cell = sheet.cell(row=row_idx, column=col_idx)
                font = Font(name=cell.font.name, size=cell.font.size)
                # 复制列宽
                if row_idx == 1:
                    col_letter = get_column_letter(col_idx)
                    new_sheet.column_dimensions[col_letter].width = sheet.column_dimensions[col_letter].width
                if col_idx == 1:
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    if value is None:
                        break
                
                # 检查当前单元格是否在合并单元格范围内
                for merged_range in merged_ranges:
                    if merged_range.min_row <= row_idx <= merged_range.max_row and \
                      merged_range.min_col <= col_idx <= merged_range.max_col:
                        in_merged_range = True
                        break
                if in_merged_range:
                    print(f'第{row_idx}行数据需要拆分.')
                    # 如果在合并单元格范围内，获取合并单元格的左上角单元格的值
                    if col_idx == 13 :
                      if row_idx >= merged_range.max_row:
                          # 统计所有的数量
                          totalNum = 0
                          merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value or 0

                          for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                              value = sheet.cell(row=r_idx, column=col_idx - 1).value or 0
                              totalNum = totalNum + float(value)

                          # 计算平均值
                          quantity = float(merged_value) / totalNum
                          print(f'开始拆分[重量]，共合并{merged_range.max_row - merged_range.min_row}行，总数量：{totalNum},总重量：{merged_value},平均值：{quantity}')

                          for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                              # 当前行的数量
                              num = sheet.cell(row=r_idx, column=col_idx - 1).value or 0
                              # 当前行的总重量
                              totalWeight = round(quantity * num, 2)
                              print(f'当前行:{r_idx},数量：{num},当前项的平均值：{quantity}, 当前行总重量：{totalWeight}')
                              new_sheet.cell(row=r_idx, column=col_idx, value=totalWeight)
                              new_sheet.cell(row=row_idx, column=col_idx).alignment = alignment
                              new_sheet.cell(row=row_idx, column=col_idx).font = font

                    else:
                        # new_row.append(cell)
                        merged_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value or 0
                        new_sheet.cell(row=row_idx, column=col_idx, value=merged_value)
                        new_sheet.cell(row=row_idx, column=col_idx).alignment = alignment
                        new_sheet.cell(row=row_idx, column=col_idx).font = font
                    
                    in_merged_range = False  # 重置标记
                else:
                    # 如果不在合并单元格范围内，直接使用当前单元格的值
                    # new_row.append(cell)
                    new_sheet.cell(row=row_idx, column=col_idx, value=cell)
                    new_sheet.cell(row=row_idx, column=col_idx).alignment = alignment
                    new_sheet.cell(row=row_idx, column=col_idx).font = font
                    in_merged_range = False  # 重置标记
            # print(row_idx,new_row)
            # new_sheet.append(new_row)

        # 复制图片
        print('复制图片', sheet._images)
        # copy_images(sheet, new_sheet)
        for image in sheet._images:
          new_sheet.add_image(image)
          print(image.anchor._from.row )	#可以获取图片的行
          print(image.anchor._from.col )	#图片的列

        file_name, file_extension = os.path.splitext(os.path.basename(file_path))
        new_file_name = f"{file_name}_拆分表{file_extension}"
        new_file_path = os.path.dirname(file_path) +'/'+ new_file_name

        print(f'原文件地址: {file_path}')
        print(f'新文件地址: {new_file_path}')
        
        new_workbook.save(new_file_path)

        # 关闭工作簿
        workbook.close()
        new_workbook.close()
    except Exception as e:
        print("处理出错：", e)
