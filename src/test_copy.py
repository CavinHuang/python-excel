from openpyxl import load_workbook, Workbook
from copy_sheet import copy_sheet

# 打开工作簿
workbook = load_workbook('./test_all.xlsx')

# 选择要处理的工作表
sheet = workbook.active

# 使用示例
tag_wb = Workbook()
tag_sheet = tag_wb.active

copy_sheet(sheet, tag_sheet)

tag_wb.save('./test_all_copy.xlsx')