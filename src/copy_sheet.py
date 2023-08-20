from copy import copy
from openpyxl.utils import get_column_letter

from openpyxl.styles import Alignment, Font

alignment = Alignment(wrapText=True, shrinkToFit=False, wrap_text=True, vertical='center', horizontal='center')

def copy_sheet(src_ws, targ_ws):
    max_row = src_ws.max_row  # 最大行数
    max_column = src_ws.max_column  # 最大列数
    w, h = 0, 0
    #复制每个单元格
    for column in range(1, max_column + 1):
        for row in range(1, max_row + 1):
            column_n = get_column_letter(column)
            i = '%s%d' % (column_n, row)  # 单元格编号
            try:
                #复制
                targ_ws[i].value = copy(src_ws[i].value)
                targ_ws[i].font = Font(src_ws[i].font.name, src_ws[i].font.size) #copy(src_ws[i].font)
                targ_ws[i].border = copy(src_ws[i].border)
                targ_ws[i].fill = copy(src_ws[i].fill)
                targ_ws[i].number_format = copy(src_ws[i].number_format)
                targ_ws[i].protection = copy(src_ws[i].protection)
                targ_ws[i].alignment = alignment #copy(src_ws[i].alignment)
            except Exception as e :
                print(e)
    #此处有坑当你获得一个列宽为13的时候实际上是这个列和前面单元格一样的宽度，并不是他真的是13
    for i in range(1, max_column + 1):
        column_letter = get_column_letter(i)
        rs = src_ws.column_dimensions[column_letter].width
        if (rs == 13):
            rs = w
        else:
            w = rs
        targ_ws.column_dimensions[column_letter].width = rs
    #复制行高，没有列宽的坑
    for i in range(1, max_row + 1):
        rs = src_ws.row_dimensions[i].height
        if rs != None:
            targ_ws.row_dimensions[i].height = rs

    wm = list(src_ws.merged_cells)  # 开始处理合并单元格
    for i in range(0, len(wm)):
        cell2 = str(wm[i]).replace('(<MergedCellRange ', '').replace('>,)', '')
        targ_ws.merge_cells(cell2)