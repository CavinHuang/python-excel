from openpyxl import load_workbook, Workbook
from PIL import Image

# 打开工作簿
workbook = load_workbook('./test_all.xlsx')

# 选择要处理的工作表
sheet = workbook.active

# 使用示例
tag_wb = Workbook()
src_sheet = sheet.active
tag_sheet = tag_wb.active

print(sheet._images)

# 假设图片在工作表的A1单元格中
# image_cell = sheet['A1']

# 获取图片对象
# image = image_cell._image

# 获取图片的二进制数据
# image_data = image.image_data

# 保存图片到磁盘
# with open('output_image.png', 'wb') as f:
#     f.write(image_data)

# # 打开并显示保存的图片
# saved_image = Image.open('output_image.png')
# saved_image.show()